import os
import openpyxl
import win32com.client as win32


# Get the current working directory
cwd = os.getcwd()

# Load the Excel workbook
workbook = openpyxl.load_workbook(os.path.join(cwd, "test_email_file.xlsx"))

# Select the sheet
sheet = workbook["Email_List"]

# Get the Outlook application object
outlook = win32.Dispatch('outlook.application')

# Iterate through the rows in the sheet
for i in range(2, sheet.max_row + 1):

    # # Get the attachment file name
    # attachment = sheet.cell(row=i, column=1).value
    # attachment_path = os.path.join(cwd, "Attachments", attachment)
    # if not os.path.exists(attachment_path):
    #     print(f"Attachment {attachment} does not exist")
    #     continue

    # Get the recipient name
    recipient_name = sheet.cell(row=i, column=2).value

    # Get the recipient email address
    recipient_email = sheet.cell(row=i, column=3).value

    # Get the CC email address
    cc_email = sheet.cell(row=i, column=4).value

    # Get the figma + dashboard links
    item_usage_link = sheet.cell(row=i, column=5).value
    item_views_link = sheet.cell(row=i, column=6).value
    old_dashboard_link = sheet.cell(row=i, column=7).value

    # Create a new email
    mail = outlook.CreateItem(0)

    # Set the recipient and CC email addresses
    mail.To = recipient_email
    mail.CC = cc_email

    # Set the email subject
    mail.Subject = f"User Testing of NEW PMcK Marketplace Item dashboards!!"

    # Set the email text
    mail.Body = f"Hey {recipient_name} – hope you’re doing well! 😊\n\nThanks again for chatting with us on the Platform McKinsey Marketplace item dashboards! We actually have reworked two dashboards into new interactive prototypes in Figma links below:\n -	Item Usage dashboard: {item_usage_link}\n -	Item Views dashboard: {item_views_link} \n  o	Item health dashboard is currently being iterated on \n\nASK: Can you please review the new prototypes at the links above and provide feedback?  \n-	NOTE: there is NO updated data so you can’t see your product, we want you to be more focused on the usability / simplicity / metrics / etc.\nThe links to current marketplace item dashboards are in the Curator Home in PMck, with the panel on the left side to change between the different dashboards: {old_dashboard_link}\n\nThanks,\nJoe Apollo"

    # Add the attachment
    # mail.Attachments.Add(attachment_path)

    # Open the email in Outlook
    mail.Display()
    
# close all opened objects
workbook.close()
